#soms is de actieve directory anders dan gewenst, wanneer dit zo is weizig deze dan via de os.chdir() command, dit is zeker niet altijd zo

#verander de directory naar de huidige dir
import os                  
os.chdir(os.path.dirname(__file__))

#hier gaan we kijken naar python mogelijkheden binnen Excel
#hiervoor gebruiken we Open Py Xl, deze kan worden gedownload via --> pip install openpyxl

#import load workbook, hiermee kunnen we excel documenten bewerken
from openpyxl import load_workbook, Workbook


#importeer het locale python document klas
import klas







#definieer het excel document
wb = load_workbook(filename=os.path.join("db", "personen.xlsx"))

#selecteer de sheet uit het excel document
ws = wb["personen"]


#maak een lege lijst voor personen
personen = []
#loop door alle rijen heen vanaf een parameter
for rij in ws.iter_rows(min_row=2):
    
    #maak een persoon aan
    persoon = klas.Persoon(rij[0].value, rij[1].value, rij[2].value, rij[3].value,)

    persoon.is_jarig()

    #voeg persoon toe aan de lijst
    personen.append(persoon)

#maak een nieuw excel document
wb = Workbook()
#selecteer de openstaande sheet
ws = wb.active

#maak voor ieder persoon in de lijst een nieuwe rij
for persoon in personen:
    ws.append([persoon.naam, persoon.gewicht, persoon.leeftijd, persoon.woonplaats])

#sla het document op
try:
    wb.save(os.path.join("db", "een_jaar_ouder.xlsx"))
    print("Document opgeslagen!")
except:
    print("Opslaan mislukt.")